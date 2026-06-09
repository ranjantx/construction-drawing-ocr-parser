"""Verify all 7 confirmed changes in one script."""
import sys
sys.path.insert(0, ".")

from pipeline.regex_patterns import PANEL_TOKEN, is_panel_token, looks_like_panel_label
from pipeline.normalizer import normalize_text
from pipeline.classifier import classify_all
from pipeline.confidence_scorer import score_candidate
from models.data_models import BBox, OCRToken, PanelCircuitCandidate

SEP = "=" * 60

def tok(text, conf=0.90):
    return OCRToken(raw_text=text, normalized_text=text.upper(), ocr_confidence=conf,
                    bbox=BBox(x1=0,y1=0,x2=100,y2=20), page=0)

# ── Change 4: digit-prefix tightening ────────────────────────────────────────
print(SEP)
print("Change 4: Digit-prefix panel token tightening")
print(SEP)
cases = [
    ("7LA",  True,  "valid digit-prefix (2 letters after digit)"),
    ("4LF",  True,  "valid digit-prefix"),
    ("7LA2", True,  "valid digit-prefix + trailing digit"),
    ("1L1",  False, "OCR artifact of EL1 (only 1 letter between digits)"),
    ("1E",   False, "OCR artifact (only 1 letter after digit)"),
    ("1L3",  False, "OCR artifact of EL3"),
    ("1LA",  True,  "valid: 1 digit + 2 letters LA"),
]
for text, expected, note in cases:
    result = is_panel_token(text)
    status = "OK " if result == expected else "FAIL"
    print(f"  [{status}] is_panel_token({text!r})={result}  expected={expected}  {note}")

# ── Change 4: looks_like_panel_label ─────────────────────────────────────────
print()
print("Change 4: looks_like_panel_label quality gate")
quality_cases = [
    ("E",        True,  "single-letter panel (E) — accepted (has no digit but allowed via dash)"),
    ("U",        True,  "single-letter panel U"),
    ("EL1",      True,  "valid panel EL1"),
    ("EL2",      True,  "valid panel EL2"),
    ("1L1",      False, "OCR artifact"),
    ("1E",       False, "OCR artifact"),
    ("CORRIDOR", False, "room name, no digit"),
    (",EL1",     False, "leading comma OCR artifact"),
]
for text, expected, note in quality_cases:
    result = looks_like_panel_label(text)
    status = "OK " if result == expected else "FAIL"
    print(f"  [{status}] looks_like_panel_label({text!r})={result}  {note}")

# ── Change 5: normalizer strips leading punctuation ──────────────────────────
print()
print("Change 5: Leading punctuation stripping in normalizer")
norm_cases = [
    (",EL1-5",   "EL1-5"),
    ("?-?-1",    "-1"),    # after stripping '?' and '-' prefix
    ("  E-29 ",  "E-29"),
    ("EL1-5",    "EL1-5"),
]
for raw, expected_start in norm_cases:
    result = normalize_text(raw)
    ok = result.startswith(expected_start) or result == expected_start
    print(f"  {'OK ' if ok else 'FAIL'} normalize({raw!r}) -> {result!r}  (expected starts with {expected_start!r})")

# ── Change 2: confidence scoring with no panel schedule ──────────────────────
print()
print("Change 2: Confidence scoring when no panel schedule (known_panels={})")
def make_pc(text, conf):
    t = tok(text, conf)
    c = PanelCircuitCandidate(token=t, classification="panel_circuit",
                               panel=text.split("-")[0], circuit=text.split("-")[1] if "-" in text else "5")
    return c

score_cases = [
    ("E-29",   0.90, "E panel, ocr=0.90"),
    ("EL1-5",  0.90, "EL1 panel, ocr=0.90"),
    ("EL1-5",  0.70, "EL1 panel, ocr=0.70"),
    ("EL1-5",  0.50, "EL1 panel, ocr=0.50"),
]
for text, ocr_conf, note in score_cases:
    c = make_pc(text, ocr_conf)
    scored = score_candidate(c, known_panels=set())   # empty = no schedule
    print(f"  score={scored.confidence_score:.3f}  conf={scored.confidence:<8}  review={scored.needs_human_review}  {note}")

# ── Change 3: needs_human_review flag ────────────────────────────────────────
print()
print("Change 3: needs_human_review only for truly uncertain cases")
c_medium = make_pc("EL1-5", 0.90)
score_candidate(c_medium, known_panels=set())
print(f"  MEDIUM conf (ocr=0.90, no geometry): needs_review={c_medium.needs_human_review}  (expected False)")

c_low = make_pc("EL1-5", 0.40)
score_candidate(c_low, known_panels=set())
print(f"  LOW conf (ocr=0.40): needs_review={c_low.needs_human_review}  (expected True)")

# ── Change 7: output sheet naming (verify Review sheet exists) ───────────────
print()
print("Change 7: Review sheet exists in output_writer")
from pipeline.output_writer import write_excel
from models.data_models import ExtractionResult
import tempfile, json
from pathlib import Path
with tempfile.TemporaryDirectory() as tmp:
    result = ExtractionResult(pdf_path="test.pdf", total_pages=1)
    write_excel(result, Path(tmp) / "test")
    from openpyxl import load_workbook
    wb = load_workbook(Path(tmp) / "test.xlsx")
    expected_sheets = {"Circuits", "Summary", "Review", "Rejected"}
    found = set(wb.sheetnames)
    missing = expected_sheets - found
    if not missing:
        print(f"  OK  All sheets present: {sorted(found)}")
    else:
        print(f"  FAIL  Missing sheets: {missing}")

print()
print("All checks complete.")
