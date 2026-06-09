"""
Token Parser CLI — Parse electrical panel/circuit tokens from JSON/XLSX/CSV files.

This CLI tool accepts tokens extracted from PDFs and enriches them with:
- Normalized text (dash standardization, case correction)
- Classification (panel, circuit, equipment, fixture, mounting height, etc.)
- Panel and circuit extraction
- Confidence scoring
- Human review flagging

Supported input formats: JSON, XLSX, CSV
Supported output formats: Same as input (with enriched columns)

Usage:
  python token_parser.py --input tokens.json --output parsed_tokens.json
  python token_parser.py --input tokens.xlsx --output parsed_tokens.xlsx --dpi 300
  python token_parser.py --input tokens.csv --output parsed_tokens.csv --verbose

Required input columns:
  - tile_id: Unique identifier for tile
  - raw_text: OCR text from token
  - bbox_x1, bbox_y1, bbox_x2, bbox_y2: Bounding box coordinates

Output columns (added/updated):
  - normalized_text: Dash standardization, case correction
  - classification: panel_circuit, equipment_tag, fixture_device_tag, etc.
  - panel: Extracted panel label (e.g., "EL1", "UL1")
  - circuit: Extracted circuit number(s) (e.g., "25", "2,4,6,8")
  - confidence: HIGH, MEDIUM, LOW, or reject
  - reason: Why classified/scored as such
  - needs_human_review: True if confidence=LOW or geometry-only match
"""

from __future__ import annotations

import argparse
import csv
import json
import logging
import sys
from pathlib import Path
from typing import Optional

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from models.data_models import BBox, OCRToken, PanelCircuitCandidate
from pipeline.classifier import classify_all
from pipeline.confidence_scorer import score_all
from pipeline.normalizer import normalize_text
from pipeline.regex_patterns import looks_like_panel_label


# ─────────────────────────────────────────────────────────────────────────────
# Logging setup
# ─────────────────────────────────────────────────────────────────────────────

def _setup_logging(verbose: bool) -> None:
    level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(
        format="%(asctime)s [%(levelname)s] %(name)s — %(message)s",
        datefmt="%H:%M:%S",
        level=level,
    )


# ─────────────────────────────────────────────────────────────────────────────
# File I/O handlers
# ─────────────────────────────────────────────────────────────────────────────

def _read_json(path: Path) -> list[dict]:
    """Read tokens from JSON file."""
    logger = logging.getLogger(__name__)
    try:
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        if not isinstance(data, list):
            data = [data]
        logger.info(f"Loaded {len(data)} tokens from {path}")
        return data
    except json.JSONDecodeError as e:
        logger.error(f"Invalid JSON in {path}: {e}")
        sys.exit(1)
    except Exception as e:
        logger.error(f"Failed to read {path}: {e}")
        sys.exit(1)


def _read_xlsx(path: Path) -> list[dict]:
    """Read tokens from Excel file."""
    logger = logging.getLogger(__name__)
    if not HAS_OPENPYXL:
        logger.error("openpyxl not installed. Install with: pip install openpyxl")
        sys.exit(1)

    try:
        wb = openpyxl.load_workbook(path)
        ws = wb.active

        # Read header from first row
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(cell.value)

        if not headers:
            logger.error(f"No headers found in {path}")
            sys.exit(1)

        # Read data rows
        rows = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            row_dict = {}
            for col_idx, header in enumerate(headers):
                if col_idx < len(row):
                    row_dict[header] = row[col_idx]
            rows.append(row_dict)

        logger.info(f"Loaded {len(rows)} tokens from {path}")
        return rows
    except Exception as e:
        logger.error(f"Failed to read {path}: {e}")
        sys.exit(1)


def _read_csv(path: Path) -> list[dict]:
    """Read tokens from CSV file."""
    logger = logging.getLogger(__name__)
    try:
        rows = []
        with open(path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            if not reader.fieldnames:
                logger.error(f"No headers in {path}")
                sys.exit(1)
            rows = list(reader)

        logger.info(f"Loaded {len(rows)} tokens from {path}")
        return rows
    except Exception as e:
        logger.error(f"Failed to read {path}: {e}")
        sys.exit(1)


def _read_input_file(path: Path) -> list[dict]:
    """Read input file (auto-detect format)."""
    logger = logging.getLogger(__name__)

    if path.suffix.lower() == '.json':
        return _read_json(path)
    elif path.suffix.lower() in ('.xlsx', '.xls'):
        return _read_xlsx(path)
    elif path.suffix.lower() == '.csv':
        return _read_csv(path)
    else:
        logger.error(f"Unsupported file format: {path.suffix}")
        logger.info("Supported formats: .json, .xlsx, .csv")
        sys.exit(1)


def _write_json(path: Path, rows: list[dict]) -> None:
    """Write tokens to JSON file."""
    logger = logging.getLogger(__name__)
    try:
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(rows, f, indent=2)
        logger.info(f"Wrote {len(rows)} tokens to {path}")
    except Exception as e:
        logger.error(f"Failed to write {path}: {e}")
        sys.exit(1)


def _write_xlsx(path: Path, rows: list[dict]) -> None:
    """Write tokens to Excel file."""
    logger = logging.getLogger(__name__)
    if not HAS_OPENPYXL:
        logger.error("openpyxl not installed. Install with: pip install openpyxl")
        sys.exit(1)

    try:
        wb = openpyxl.Workbook()
        ws = wb.active

        if not rows:
            logger.warning("No rows to write")
            return

        # Get headers from first row
        headers = list(rows[0].keys())

        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        # Write data
        for row_idx, row_data in enumerate(rows, start=2):
            for col_idx, header in enumerate(headers, start=1):
                value = row_data.get(header)
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-fit columns
        for col_idx, header in enumerate(headers, start=1):
            max_len = len(str(header))
            for row_data in rows:
                cell_val = str(row_data.get(header, ""))
                max_len = max(max_len, len(cell_val))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 2, 50)

        wb.save(path)
        logger.info(f"Wrote {len(rows)} tokens to {path}")
    except Exception as e:
        logger.error(f"Failed to write {path}: {e}")
        sys.exit(1)


def _write_csv(path: Path, rows: list[dict]) -> None:
    """Write tokens to CSV file."""
    logger = logging.getLogger(__name__)
    try:
        if not rows:
            logger.warning("No rows to write")
            return

        headers = list(rows[0].keys())
        with open(path, 'w', encoding='utf-8', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            writer.writerows(rows)

        logger.info(f"Wrote {len(rows)} tokens to {path}")
    except Exception as e:
        logger.error(f"Failed to write {path}: {e}")
        sys.exit(1)


def _write_output_file(path: Path, rows: list[dict]) -> None:
    """Write output file (auto-detect format)."""
    logger = logging.getLogger(__name__)

    if path.suffix.lower() == '.json':
        _write_json(path, rows)
    elif path.suffix.lower() in ('.xlsx', '.xls'):
        _write_xlsx(path, rows)
    elif path.suffix.lower() == '.csv':
        _write_csv(path, rows)
    else:
        logger.error(f"Unsupported output format: {path.suffix}")
        logger.info("Supported formats: .json, .xlsx, .csv")
        sys.exit(1)


# ─────────────────────────────────────────────────────────────────────────────
# Token validation and conversion
# ─────────────────────────────────────────────────────────────────────────────

def _validate_input_columns(rows: list[dict]) -> None:
    """Validate required input columns are present."""
    logger = logging.getLogger(__name__)

    required_cols = {'tile_id', 'raw_text', 'bbox_x1', 'bbox_y1', 'bbox_x2', 'bbox_y2'}

    if not rows:
        logger.error("Input file has no rows")
        sys.exit(1)

    available_cols = set(rows[0].keys())
    missing_cols = required_cols - available_cols

    if missing_cols:
        logger.error(f"Missing required columns: {', '.join(sorted(missing_cols))}")
        logger.info(f"Available columns: {', '.join(sorted(available_cols))}")
        sys.exit(1)

    logger.info(f"✓ All required columns present: {', '.join(sorted(required_cols))}")


def _convert_to_ocr_tokens(rows: list[dict], dpi: int = 300) -> list[OCRToken]:
    """Convert input rows to OCRToken objects."""
    logger = logging.getLogger(__name__)

    tokens = []
    px_per_pt = dpi / 72.0

    for row_idx, row in enumerate(rows):
        try:
            # Extract bbox coordinates (in points, convert to pixels)
            x1_pt = float(row.get('bbox_x1', 0))
            y1_pt = float(row.get('bbox_y1', 0))
            x2_pt = float(row.get('bbox_x2', 0))
            y2_pt = float(row.get('bbox_y2', 0))

            x1_px = x1_pt * px_per_pt
            y1_px = y1_pt * px_per_pt
            x2_px = x2_pt * px_per_pt
            y2_px = y2_pt * px_per_pt

            bbox = BBox(x1=x1_px, y1=y1_px, x2=x2_px, y2=y2_px)

            token = OCRToken(
                raw_text=str(row.get('raw_text', '')).strip(),
                normalized_text='',  # Will be filled by normalizer
                ocr_confidence=float(row.get('ocr_confidence', 1.0)),
                bbox=bbox,
                page=int(row.get('page', 0)),
                tile_id=str(row.get('tile_id', f'tile_{row_idx}'))
            )
            tokens.append(token)
        except (ValueError, KeyError) as e:
            logger.warning(f"Row {row_idx}: Failed to parse bbox coordinates: {e}")
            continue

    logger.info(f"Converted {len(tokens)} tokens from {len(rows)} rows")
    return tokens


# ─────────────────────────────────────────────────────────────────────────────
# Token enrichment (parse tokens using existing pipeline)
# ─────────────────────────────────────────────────────────────────────────────

def _enrich_tokens(tokens: list[OCRToken], known_panels: set[str] | None = None) -> list[PanelCircuitCandidate]:
    """Enrich tokens with classification, normalization, and scoring."""
    logger = logging.getLogger(__name__)

    if known_panels is None:
        known_panels = set()

    logger.info("Step 1: Normalizing text...")
    for token in tokens:
        token.normalized_text = normalize_text(token.raw_text, known_panels)

    logger.info("Step 2: Classifying tokens...")
    candidates = classify_all(tokens, known_panels=known_panels)

    logger.info("Step 3: Scoring candidates...")
    candidates = score_all(candidates, known_panels)

    return candidates


def _candidates_to_output_rows(candidates: list[PanelCircuitCandidate],
                                input_rows: list[dict]) -> list[dict]:
    """Convert enriched candidates back to output format."""
    # Create mapping from token id to input row (for preserving extra columns)
    input_map = {}
    for row in input_rows:
        # Use tile_id + raw_text as key
        key = (str(row.get('tile_id', '')), str(row.get('raw_text', '')))
        input_map[key] = row

    output_rows = []
    for candidate in candidates:
        token = candidate.token

        # Preserve original row data
        output_row = input_map.get((token.tile_id, token.raw_text), {}).copy()

        # Update with enriched columns
        output_row.update({
            'tile_id': token.tile_id,
            'raw_text': token.raw_text,
            'normalized_text': token.normalized_text,
            'classification': candidate.classification,
            'panel': candidate.panel or '',
            'circuit': candidate.circuit or '',
            'confidence': candidate.confidence,
            'reason': candidate.reason,
            'needs_human_review': candidate.needs_human_review,
        })

        output_rows.append(output_row)

    return output_rows


# ─────────────────────────────────────────────────────────────────────────────
# Main CLI
# ─────────────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        prog='token_parser',
        description='Parse electrical panel/circuit tokens from JSON/XLSX/CSV files.',
        epilog='''
Examples:
  # Parse JSON tokens
  python token_parser.py --input tokens.json --output parsed.json

  # Parse Excel with custom DPI
  python token_parser.py --input tokens.xlsx --output parsed.xlsx --dpi 300

  # Parse CSV with verbose logging
  python token_parser.py --input tokens.csv --output parsed.csv --verbose

  # Parse with known panels
  python token_parser.py --input tokens.json --output parsed.json --known-panels EL1,EL2,UL1

Supported input formats: JSON, XLSX, CSV
Supported output formats: Same as input (with enriched columns)

Required input columns:
  - tile_id: Unique identifier for token
  - raw_text: OCR text from PDF
  - bbox_x1, bbox_y1, bbox_x2, bbox_y2: Bounding box in PDF points

Output columns (added/updated):
  - normalized_text: Standardized text (dashes, case)
  - classification: panel_circuit, equipment_tag, fixture_device_tag, etc.
  - panel: Extracted panel label (e.g., "EL1", "UL1")
  - circuit: Extracted circuit numbers (e.g., "25" or "2,4,6")
  - confidence: HIGH, MEDIUM, LOW, or reject
  - reason: Classification and scoring rationale
  - needs_human_review: True if LOW confidence or geometry-only match
        ''',
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )

    parser.add_argument(
        '--input', '-i',
        type=Path,
        required=True,
        help='Input file (JSON/XLSX/CSV with tokens)',
    )

    parser.add_argument(
        '--output', '-o',
        type=Path,
        required=True,
        help='Output file (same format as input)',
    )

    parser.add_argument(
        '--dpi',
        type=int,
        default=300,
        help='DPI for PDF coordinate conversion (default: 300)',
    )

    parser.add_argument(
        '--known-panels',
        type=str,
        default='',
        help='Comma-separated list of known panel labels (e.g., "EL1,EL2,UL1")',
    )

    parser.add_argument(
        '--verbose', '-v',
        action='store_true',
        help='Enable debug logging',
    )

    args = parser.parse_args()

    # Setup logging
    _setup_logging(args.verbose)
    logger = logging.getLogger(__name__)

    logger.info("═" * 80)
    logger.info("Token Parser — Parse electrical panel/circuit tokens")
    logger.info("═" * 80)

    # Validate input file exists
    if not args.input.exists():
        logger.error(f"Input file not found: {args.input}")
        sys.exit(1)

    # Create output directory if needed
    args.output.parent.mkdir(parents=True, exist_ok=True)

    # Parse known panels
    known_panels = set()
    if args.known_panels:
        known_panels = set(panel.strip().upper() for panel in args.known_panels.split(',') if panel.strip())
        logger.info(f"Known panels: {', '.join(sorted(known_panels))}")

    logger.info(f"Input file: {args.input}")
    logger.info(f"Output file: {args.output}")
    logger.info(f"DPI: {args.dpi}")
    logger.info("─" * 80)

    # Read input file
    logger.info("Reading input file...")
    input_rows = _read_input_file(args.input)

    # Validate required columns
    logger.info("Validating columns...")
    _validate_input_columns(input_rows)

    # Convert to OCR tokens
    logger.info("Converting to OCR tokens...")
    tokens = _convert_to_ocr_tokens(input_rows, dpi=args.dpi)

    if not tokens:
        logger.error("No valid tokens found in input file")
        sys.exit(1)

    # Enrich tokens
    logger.info("─" * 80)
    logger.info("Enriching tokens (normalize, classify, score)...")
    candidates = _enrich_tokens(tokens, known_panels=known_panels)

    # Convert back to output format
    logger.info("Converting to output format...")
    output_rows = _candidates_to_output_rows(candidates, input_rows)

    # Write output file
    logger.info("Writing output file...")
    _write_output_file(args.output, output_rows)

    # Summary statistics
    logger.info("─" * 80)
    panel_circuits = [c for c in candidates if c.classification == 'panel_circuit']
    high_conf = [c for c in candidates if c.confidence == 'high']
    medium_conf = [c for c in candidates if c.confidence == 'medium']
    low_conf = [c for c in candidates if c.confidence == 'low']
    rejected = [c for c in candidates if c.confidence == 'reject']
    review = [c for c in candidates if c.needs_human_review]

    logger.info("═" * 80)
    logger.info("RESULTS SUMMARY")
    logger.info("═" * 80)
    logger.info(f"Total tokens processed: {len(candidates)}")
    logger.info(f"Panel-circuit matches: {len(panel_circuits)}")
    logger.info(f"  HIGH confidence: {len(high_conf)}")
    logger.info(f"  MEDIUM confidence: {len(medium_conf)}")
    logger.info(f"  LOW confidence: {len(low_conf)}")
    logger.info(f"  REJECTED: {len(rejected)}")
    logger.info(f"Flagged for human review: {len(review)}")
    logger.info("═" * 80)
    logger.info(f"✓ Output written to: {args.output}")


if __name__ == '__main__':
    main()
