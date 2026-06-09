"""
Write extraction results to Excel (3 sheets) and JSON.

Excel sheets:
  Circuits  — all 14 columns, colour-coded by confidence
  Summary   — pivot by panel: circuit count + confidence distribution
  Rejected  — rows where confidence=reject or needs_human_review=True

JSON: flat list of all candidate rows.
"""

from __future__ import annotations

import json
import logging
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from models.data_models import ExtractionResult

logger = logging.getLogger(__name__)

# Confidence → fill colour (ARGB)
_FILL = {
    "high":   PatternFill("solid", fgColor="C6EFCE"),  # green
    "medium": PatternFill("solid", fgColor="FFEB9C"),  # yellow
    "low":    PatternFill("solid", fgColor="FFCC99"),  # orange
    "reject": PatternFill("solid", fgColor="FFC7CE"),  # red
}

_COLUMNS = [
    "page", "tile_id", "raw_text", "normalized_text", "classification",
    "panel", "circuit", "confidence", "bbox_x1", "bbox_y1", "bbox_x2", "bbox_y2",
    "reason", "needs_human_review",
]

_HEADER_FILL = PatternFill("solid", fgColor="4472C4")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def _auto_fit(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 60)


def _write_header(ws, columns: list[str]) -> None:
    for i, col in enumerate(columns, 1):
        cell = ws.cell(row=1, column=i, value=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def write_excel(result: ExtractionResult, output_path: Path) -> None:
    """
    Write four Excel sheets:

    Circuits  — all panel_circuit matches (green=high, yellow=medium, orange=low)
    Summary   — pivot: unique panels with circuit counts and confidence breakdown
    Review    — panel_circuit rows flagged needs_human_review=True (verify these)
                These are VALID extractions but with low OCR confidence or
                geometry-only association. They also appear in Circuits.
    Rejected  — rows where confidence=reject (failed validation, invalid circuit,
                or unrelated annotation). These are NOT valid panel-circuit pairs.
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)

    rows = [c.to_row() for c in result.candidates]
    df_all = pd.DataFrame(rows, columns=_COLUMNS) if rows else pd.DataFrame(columns=_COLUMNS)

    df_circuits = df_all[df_all["classification"] == "panel_circuit"].copy()

    # Review: valid panel_circuit rows that need a human eye
    df_review = df_circuits[df_circuits["needs_human_review"] == True].copy()

    # Rejected: truly invalid — failed validation, wrong classification
    df_rejected = df_all[df_all["confidence"] == "reject"].copy()

    # Summary pivot
    if not df_circuits.empty:
        summary_rows = (
            df_circuits.groupby("panel")
            .agg(
                total_circuits=("circuit", "count"),
                high=("confidence", lambda s: (s == "high").sum()),
                medium=("confidence", lambda s: (s == "medium").sum()),
                low=("confidence", lambda s: (s == "low").sum()),
                needs_review=("needs_human_review", "sum"),
            )
            .reset_index()
            .sort_values("total_circuits", ascending=False)
        )
    else:
        summary_rows = pd.DataFrame(
            columns=["panel", "total_circuits", "high", "medium", "low", "needs_review"]
        )

    excel_path = output_path.with_suffix(".xlsx")
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df_circuits.to_excel(writer, sheet_name="Circuits", index=False)
        summary_rows.to_excel(writer, sheet_name="Summary", index=False)
        df_review.to_excel(writer, sheet_name="Review", index=False)
        df_rejected.to_excel(writer, sheet_name="Rejected", index=False)

    # Post-process: colour rows, style headers, auto-fit columns
    wb = load_workbook(excel_path)

    # Circuits sheet — colour by confidence
    ws = wb["Circuits"]
    _write_header(ws, _COLUMNS)
    conf_col = _COLUMNS.index("confidence") + 1
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        conf_val = row[conf_col - 1].value
        fill = _FILL.get(str(conf_val), _FILL["reject"])
        for cell in row:
            cell.fill = fill
    _auto_fit(ws)

    # Review sheet — all orange (needs human eye)
    ws_rev = wb["Review"]
    for row in ws_rev.iter_rows(min_row=2, max_row=ws_rev.max_row):
        for cell in row:
            cell.fill = _FILL["low"]
    _auto_fit(ws_rev)

    for sheet_name in ("Summary", "Rejected"):
        _auto_fit(wb[sheet_name])

    wb.save(excel_path)
    logger.info("Excel written: %s  (%d circuits, %d review, %d rejected)",
                excel_path, len(df_circuits), len(df_review), len(df_rejected))


def write_json(result: ExtractionResult, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    json_path = output_path.with_suffix(".json")
    payload = {
        "pdf_path": result.pdf_path,
        "total_pages": result.total_pages,
        "known_panels": sorted(result.known_panels),
        "candidates": [c.to_row() for c in result.candidates],
    }
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, default=str)
    logger.info("JSON written: %s", json_path)


def write_outputs(result: ExtractionResult, output_dir: str | Path) -> None:
    """
    Always write to a new timestamped file — never overwrites a previous run,
    never collides with an open Excel file.

    Output names:
      circuits_YYYYMMDD_HHMMSS.xlsx
      circuits_YYYYMMDD_HHMMSS.json
    """
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    stem = out_dir / f"circuits_{ts}"
    write_excel(result, stem)   # appends .xlsx
    write_json(result,  stem)   # appends .json
    logger.info("Run timestamp: %s", ts)
