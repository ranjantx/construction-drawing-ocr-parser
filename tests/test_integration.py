"""
Integration tests — run the full pipeline on a sample PDF.
Tests are skipped if the fixture PDF does not exist.

All tests share a single session-scoped pipeline run (see conftest.pipeline_output_dir)
so OCR only runs once for the whole session.

Place your PDF at:
  tests/fixtures/sample_panel_labels.pdf
"""

from __future__ import annotations

import pytest
from pathlib import Path

from tests.conftest import SAMPLE_PDF

pytestmark = pytest.mark.skipif(
    not SAMPLE_PDF.exists(),
    reason=f"Integration fixture PDF not found: {SAMPLE_PDF}",
)


def test_pipeline_runs_without_error(pipeline_output_dir):
    """Full pipeline must complete and produce output files."""
    assert (pipeline_output_dir / "circuits.xlsx").exists(), "Excel output missing"
    assert (pipeline_output_dir / "circuits.json").exists(), "JSON output missing"


def test_output_has_required_columns(pipeline_output_dir):
    """circuits.xlsx Circuits sheet must have all 14 required columns."""
    import pandas as pd
    df = pd.read_excel(pipeline_output_dir / "circuits.xlsx", sheet_name="Circuits")
    required = [
        "page", "tile_id", "raw_text", "normalized_text", "classification",
        "panel", "circuit", "confidence", "bbox_x1", "bbox_y1", "bbox_x2", "bbox_y2",
        "reason", "needs_human_review",
    ]
    for col in required:
        assert col in df.columns, f"Missing column: {col}"


def test_mounting_heights_not_in_circuits(pipeline_output_dir):
    """Mounting height tokens must never appear in the Circuits sheet as panel_circuit."""
    import pandas as pd
    df = pd.read_excel(pipeline_output_dir / "circuits.xlsx", sheet_name="Circuits")
    mounting = df[df["classification"] == "mounting_height"]
    assert mounting.empty, f"Mounting heights found in Circuits sheet: {mounting['raw_text'].tolist()}"


def test_no_circuits_above_84(pipeline_output_dir):
    """No extracted circuit number may exceed 84."""
    import pandas as pd
    df = pd.read_excel(pipeline_output_dir / "circuits.xlsx", sheet_name="Circuits")
    pc = df[df["classification"] == "panel_circuit"]
    for _, row in pc.iterrows():
        for c in str(row["circuit"]).split(","):
            c = c.strip()
            if c.isdigit():
                assert int(c) <= 84, f"Circuit {c} > 84 in panel_circuit row"


def test_excel_has_four_sheets(pipeline_output_dir):
    """Excel file must have Circuits, Summary, Review, and Rejected sheets."""
    from openpyxl import load_workbook
    files = list((pipeline_output_dir).glob("*.xlsx"))
    assert files, "No xlsx file found"
    wb = load_workbook(files[0])
    assert "Circuits"  in wb.sheetnames
    assert "Summary"   in wb.sheetnames
    assert "Review"    in wb.sheetnames
    assert "Rejected"  in wb.sheetnames


def test_json_structure(pipeline_output_dir):
    """JSON output must have required top-level keys."""
    import json
    with open(pipeline_output_dir / "circuits.json") as f:
        data = json.load(f)
    assert "pdf_path" in data
    assert "total_pages" in data
    assert "candidates" in data
    assert "known_panels" in data
    assert isinstance(data["candidates"], list)


def test_at_least_one_panel_circuit_found(pipeline_output_dir):
    """At least one panel-circuit pair must be extracted from the fixture PDF."""
    import pandas as pd
    df = pd.read_excel(pipeline_output_dir / "circuits.xlsx", sheet_name="Circuits")
    pc = df[df["classification"] == "panel_circuit"]
    assert len(pc) >= 1, "No panel-circuit pairs found in output"


def test_known_panels_in_json(pipeline_output_dir):
    """JSON must include at least one discovered panel label."""
    import json
    with open(pipeline_output_dir / "circuits.json") as f:
        data = json.load(f)
    assert len(data["known_panels"]) >= 1, "No known panels discovered"
